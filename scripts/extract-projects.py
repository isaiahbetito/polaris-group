#!/usr/bin/env python3
"""
Extract per-project data from the client's "SITO POLARIS HR SRL.xlsx" into
data/projects.json (one entry per plan sheet). Then run gen-projects.mjs to
rebuild the Projects section on numeri.html.

Usage:
  python3 scripts/extract-projects.py "/path/to/SITO POLARIS HR SRL.xlsx"
  node scripts/gen-projects.mjs

Requires: openpyxl  (pip install openpyxl)
The .xlsx itself is NOT committed; data/projects.json is the committed source of truth.
"""
import sys, os, json, re
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser("~/Downloads/SITO POLARIS HR SRL.xlsx")
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "projects.json")

wb = openpyxl.load_workbook(XLSX, data_only=True)
projects = []
for name in wb.sheetnames:
    if name == "SITO":
        continue
    rows = list(wb[name].iter_rows(values_only=True))
    if len(rows) < 2:
        continue
    year, code, fund, corsi, ore, az, part = (list(rows[1]) + [None] * 7)[:7]
    num = lambda x: int(x) if isinstance(x, (int, float)) else 0
    titles = []
    for r in rows[3:]:  # course-detail rows
        r = list(r)
        if len(r) > 4 and isinstance(r[4], str) and r[1]:
            t = re.sub(r"\s+", " ", r[4].strip())
            if t and t not in titles:
                titles.append(t)
    projects.append({
        "code": str(code).strip() if code else name,
        "fund": str(fund).strip() if fund else "",
        "year": int(year) if isinstance(year, (int, float)) else None,
        "corsi": num(corsi), "ore": num(ore), "aziende": num(az), "partecipanti": num(part),
        "temi": len(titles), "titoli": titles,
    })

projects.sort(key=lambda p: p["code"], reverse=True)
os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump({"source": os.path.basename(XLSX), "count": len(projects), "projects": projects},
              f, ensure_ascii=False, indent=2)
print(f"wrote {OUT} with {len(projects)} projects")
